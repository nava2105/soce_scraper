from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.writer.excel import save_workbook
from openpyxl import load_workbook
import pandas as pd
import os

OUTPUT_FOLDER = "outputs"

os.makedirs(OUTPUT_FOLDER, exist_ok=True)

def export_excel_process_inf(records, start_date, end_date):
    df = pd.DataFrame(records)

    # Rename columns
    df.rename(columns={
        "nu": "Nro.",
        "n": "Nro.Factura",
        "f": "Fecha de emisión de la factura",
        "c": "CPC",
        "d": "Descripción CPC",
        "r": "Razón Social",
        "co": "Objeto de Compra",
        "ca": "Cantidad",
        "t": "Costo U.",
        "j": "Justificativo",
        "ct": "Tipo de Compra",
        "nom": "Responsable de Asuntos Administrativos"
    }, inplace=True)

    # Delete the "no" and "i" columns if they exist
    df.drop(columns=["no", "i"], errors='ignore', inplace=True)

    # Convert "Quantity" and "U. Cost" to numeric type
    df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors='coerce')
    df["Costo U."] = pd.to_numeric(df["Costo U."], errors='coerce')

    # Create the "Value" column by multiplying "Quantity" by "U. Cost".
    df["Valor"] = df["Cantidad"] * df["Costo U."]

    # Show the first records
    print(df.head())
    excel_filename = f"INF-{start_date}-{end_date}.xlsx"
    output_file_path = os.path.join(OUTPUT_FOLDER, excel_filename)

    # Save to Excel file
    df.to_excel(output_file_path, index=False)

    return output_file_path

def export_excel_process_pro(records, start_date, end_date):
    # Convert to a Pandas DataFrame
    df = pd.DataFrame(records)

    # Rename columns
    df.rename(columns={
        "c": "Código",
        "r": "Entidad Contratante",
        "d": "Objeto del Proceso",
        "s": "Provincia/Cantón",
        "g": "Estado del Proceso",
        "p": "Presupuesto Referencial Total(sin iva)",
        "co": "Objeto de Compra",
        "f": "Fecha de Publicación",
        "i": "Link"
    }, inplace=True)

    # Delete the "no" and "i" columns if they exist
    df.drop(columns=["j", "z", "t", "u", "v", "e"], errors='ignore', inplace=True)

    # Completar la URL en la columna 'Link' sin modificar la DataFrame
    df["Link"] = df["Link"].apply(lambda
                                      x: f"https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/PC/informacionProcesoContratacion2.cpe?idSoliCompra={x}" if pd.notna(
        x) else "")

    # Guardar en Excel con hipervínculos
    excel_filename = f"PRO-{start_date}-{end_date}.xlsx"
    output_file_path = os.path.join(OUTPUT_FOLDER, excel_filename)
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())

    for row in df.itertuples(index=False):
        excel_row = list(row)
        ws.append(excel_row)

    # Aplicar hipervínculos en la última columna
    for i, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=len(df.columns), max_col=len(df.columns))):
        cell = row[0]
        if cell.value:
            cell.hyperlink = cell.value  # Mantener la URL como hipervínculo
            cell.value = "Ver Detalle"  # Mostrar solo "Ver Detalle" en Excel
            cell.font = Font(color="0000FF", underline="single")  # Estilo azul y subrayado

    wb.save(output_file_path)

    return output_file_path

def export_excel_process_reg(records, start_date, end_date):
    # Convert to a Pandas DataFrame
    df = pd.DataFrame(records)

    # Rename columns
    df.rename(columns={
        "c": "Código",
        "r": "Entidad Contratante",
        "d": "Objeto del Proceso",
        "s": "Provincia/Cantón",
        "g": "Estado del Proceso",
        "p": "Presupuesto Referencial Total(sin iva)",
        "f": "Fecha de Publicación",
        "i": "Link"
    }, inplace=True)

    # Delete the "no" and "i" columns if they exist
    df.drop(columns=["j", "z", "t", "u", "v", "e"], errors='ignore', inplace=True)

    # Completar la URL en la columna 'Link' sin modificar la DataFrame
    df["Link"] = df["Link"].apply(lambda
                                      x: f"https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/PC/informacionProcesoContratacion2.cpe?idSoliCompra={x}" if pd.notna(
        x) else "")

    # Guardar en Excel con hipervínculos
    excel_filename = f"REG-{start_date}-{end_date}.xlsx"
    output_file_path = os.path.join(OUTPUT_FOLDER, excel_filename)
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())

    for row in df.itertuples(index=False):
        excel_row = list(row)
        ws.append(excel_row)

    # Aplicar hipervínculos en la última columna
    for i, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=len(df.columns), max_col=len(df.columns))):
        cell = row[0]
        if cell.value:
            cell.hyperlink = cell.value  # Mantener la URL como hipervínculo
            cell.value = "Ver Detalle"  # Mostrar solo "Ver Detalle" en Excel
            cell.font = Font(color="0000FF", underline="single")  # Estilo azul y subrayado

    wb.save(output_file_path)

    return output_file_path

def export_excel_process_pre(records, start_date, end_date):
    # Convert to a Pandas DataFrame
    df = pd.DataFrame(records)

    # Rename columns
    df.rename(columns={
        "c": "Código",
        "r": "Entidad Contratante",
        "d": "Objeto del Proceso",
        "s": "Provincia/Cantón",
        "g": "Estado del Proceso",
        "p": "Presupuesto Referencial Total(sin iva)",
        "f": "Fecha de Publicación",
        "i": "Link"
    }, inplace=True)

    # Delete the "no" and "i" columns if they exist
    df.drop(columns=["j", "z", "t", "u", "v", "e"], errors='ignore', inplace=True)

    # Completar la URL en la columna 'Link' sin modificar la DataFrame
    df["Link"] = df["Link"].apply(lambda
                                      x: f"https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/PC/informacionProcesoContratacion2.cpe?idSoliCompra={x}" if pd.notna(
        x) else "")

    # Guardar en Excel con hipervínculos
    excel_filename = f"PRE-{start_date}-{end_date}.xlsx"
    output_file_path = os.path.join(OUTPUT_FOLDER, excel_filename)
    wb = Workbook()
    ws = wb.active
    ws.append(df.columns.tolist())

    for row in df.itertuples(index=False):
        excel_row = list(row)
        ws.append(excel_row)

    # Aplicar hipervínculos en la última columna
    for i, row in enumerate(
            ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=len(df.columns), max_col=len(df.columns))):
        cell = row[0]
        if cell.value:
            cell.hyperlink = cell.value  # Mantener la URL como hipervínculo
            cell.value = "Ver Detalle"  # Mostrar solo "Ver Detalle" en Excel
            cell.font = Font(color="0000FF", underline="single")  # Estilo azul y subrayado

    wb.save(output_file_path)

    return output_file_path

def load_hyperlinks_from_excel(file_path):
    workbook = load_workbook(file_path)
    sheet = workbook.active
    hyperlinks = []
    for row in sheet.iter_rows():
        for cell in row:
            if cell.hyperlink:
                hyperlinks.append(cell.hyperlink.target)
    return list(set(hyperlinks))  # Remove duplicates

def save_dataframe_to_excel(df, filename):
    output_file_path = os.path.join("outputs", f"{filename.replace('.xlsx', '')}-ADM.xlsx")
    
    # Save DataFrame to Excel with openpyxl engine
    df.to_excel(output_file_path, index=False, engine='openpyxl')
    
    # Load the workbook to add hyperlinks
    workbook = load_workbook(output_file_path)
    sheet = workbook.active
    
    # Find the Link column
    link_col = None
    for idx, col in enumerate(df.columns, 1):
        if col == "Link":
            link_col = idx
            break
    
    if link_col:
        # Format hyperlinks
        for row in range(2, sheet.max_row + 1):
            cell = sheet.cell(row=row, column=link_col)
            if cell.value:
                cell.hyperlink = Hyperlink(ref=cell.coordinate, 
                                         target=str(cell.value),
                                         display="Ver Detalle")
                cell.value = "Ver Detalle"
                cell.font = Font(color="0000FF", underline="single")
    
    # Save and close
    workbook.save(output_file_path)
    workbook.close()
    
    return output_file_path
