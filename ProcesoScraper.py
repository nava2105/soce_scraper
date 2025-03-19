import requests
import json
import pandas as pd
import re
import time
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

# URL we want to scrape from
url = "https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/servicio/interfazWeb.php"

# Headers to be sent with the request
headers = {
    "Accept": "text/javascript, text/html, application/xml, text/xml, */*",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "X-Requested-With": "XMLHttpRequest",
    "Origin": "https://www.compraspublicas.gob.ec",
    "Referer": "https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/IC/buscarInfima.cpe",
}

# Function to make the request and process the data
def fetch_data(start_date, end_date):
    # Data to be sent in the POST request
    data = {
        "__class": "SolicitudCompra",
        "__action": "buscarProcesoxEntidad",
        "txtEntidadContratante": "CORPORACION ELECTRICA DEL ECUADOR CELEC EP.",
        "cmbEntidad": "238940",
        "txtNroFactInfima": "",
        "txtIdProducto": "",
        "txtObsInfima": "",
        "txtCodTipoCompraTab": "176",
        "txtCodTipoCompra": "",
        "txtMes": "0",
        "txtAnio": "0",
        "f_inicio": start_date,
        "f_fin": end_date,
        "count": "521",
        "paginaActual": "0",
        "estado": "",
        "trx": "",
        "_": ""
    }

    all_data = []  # List to store all data

    # Iterate over the pages
    pagina_actual = 0
    while True:
        data["paginaActual"] = str(pagina_actual)
        response = requests.post(url, headers=headers, data=data)

        # Verify that the request was successful
        if response.status_code == 200:
            headers_json = dict(response.headers)

            if 'X-JSON' in headers_json:
                try:
                    normalized_data = re.sub(r'[()]', '', headers_json['X-JSON'])
                    json_data = json.loads(normalized_data)
                    print(pagina_actual, ": ", json_data)

                    if not json_data:  # If there is no more data, exit the loop.
                        break

                    all_data.extend(json_data)  # Add data to the list
                    pagina_actual += 20  # Go to next page

                except json.JSONDecodeError:
                    print("Error: The content of 'X-JSON' is not a valid JSON.")
                    break
            else:
                print("The 'X-JSON' header was not found.")
                break
        else:
            print(f"Error in the application: {response.status_code}")
            break

        # Wait some time before the next request to avoid blockages.
        time.sleep(1)  # Wait 1 second

    return all_data

# Define date range
start_date = "2024-09-19"
end_date = "2025-03-19"

print(f"Fetching data from {start_date} to {end_date}...")
records = fetch_data(start_date, end_date)

# Convert to a Pandas DataFrame
df = pd.DataFrame(records)

# Rename columns
df.rename(columns={
    "c": "Código",
    "r": "Entidad Contratante",
    "d": "Objeto del Proceso",
    "s": "Provincia/Cantón",
    "g": "Estado del Proceso",
    "p": "Presupuesto Referencial Total(sin iva)",
    "co": "Objeto de Compra",
    "f": "Fecha de Publicación",
    "i": "Link"
}, inplace=True)

# Delete the "no" and "i" columns if they exist
df.drop(columns=["j", "z", "t", "u", "v", "e"], errors='ignore', inplace=True)

# Completar la URL en la columna 'Link' sin modificar la DataFrame
df["Link"] = df["Link"].apply(lambda x: f"https://www.compraspublicas.gob.ec/ProcesoContratacion/compras/PC/informacionProcesoContratacion2.cpe?idSoliCompra={x}" if pd.notna(x) else "")

# Guardar en Excel con hipervínculos
excel_filename = f"PRO-{start_date}-{end_date}.xlsx"
wb = Workbook()
ws = wb.active
ws.append(df.columns.tolist())

for row in df.itertuples(index=False):
    excel_row = list(row)
    ws.append(excel_row)

# Aplicar hipervínculos en la última columna
for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=len(df.columns), max_col=len(df.columns))):
    cell = row[0]
    if cell.value:
        cell.hyperlink = cell.value  # Mantener la URL como hipervínculo
        cell.value = "Ver Detalle"  # Mostrar solo "Ver Detalle" en Excel
        cell.font = Font(color="0000FF", underline="single")  # Estilo azul y subrayado

wb.save(excel_filename)

print("Data successfully exported to Excel.")